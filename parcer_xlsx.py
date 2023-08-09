import os.path

from openpyxl import load_workbook, Workbook
from openpyxl.styles import PatternFill, Border, Side
from openpyxl.utils.cell import get_column_letter, column_index_from_string


def ParcerXlsxData(file1="Docs/--2023 (6).xlsx", file2="Docs/Оценки.xlsx", spec="Информационные системы и программирование"):

    wb_file1 = load_workbook(file1, read_only=True, data_only=True)
    wb_file2 = load_workbook(file2, read_only=True, data_only=True)
    ws_f1 = wb_file1.active
    ws_f2 = wb_file2.active

    new_file = Workbook()
    ws = new_file.active

    # Парсинг файла с оценками
    FIO = {}
    Subject = []
    Subjetc_hour = []

    for column in range(4, ws_f2.max_column + 1):
        Subject.append(ws_f2.cell(row=5, column=column).value)
        Subjetc_hour.append(ws_f2.cell(row=4, column=column).value)

    for row in range(6, ws_f2.max_row + 1):
        fio = ws_f2.cell(row=row, column=1).value
        eval = []
        for column in range(4, ws_f2.max_column + 1):
            eval.append(ws_f2.cell(row=row, column=column).value)
        eval.append(eval.pop(Subject.index('ВСЕГО часов теоретического обучения:')))
        eval.append(eval.pop(Subject.index('в том числе аудиторных часов:')))
        FIO[fio] = eval

    Subjetc_hour.append(Subjetc_hour.pop(Subject.index('ВСЕГО часов теоретического обучения:')))
    Subjetc_hour.append(Subjetc_hour.pop(Subject.index('в том числе аудиторных часов:')))

    Subject.append(Subject.pop(Subject.index('ВСЕГО часов теоретического обучения:')))
    Subject.append(Subject.pop(Subject.index('в том числе аудиторных часов:')))

    # парсинг файла с данными о студентах
    data_student = []
    for row in range(2, ws_f1.max_row + 1):
        if ws_f1.cell(row=row, column=column_index_from_string('S')).value is None:
            break
        fio = f"{ws_f1.cell(row=row, column=column_index_from_string('S')).value} " \
              f"{ws_f1.cell(row=row, column=column_index_from_string('T')).value} " \
              f"{ws_f1.cell(row=row, column=column_index_from_string('U')).value}"
        if fio in FIO.keys():
            student = {
                'surname': ws_f1.cell(row=row, column=column_index_from_string('S')).value,
                'name': ws_f1.cell(row=row, column=column_index_from_string('T')).value,
                'patronymic': ws_f1.cell(row=row, column=column_index_from_string('U')).value,
                'birthday': ws_f1.cell(row=row, column=column_index_from_string('V')).value,
                'year_references': ws_f1.cell(row=row, column=column_index_from_string('Q')).value,
                'speciality_code': ws_f1.cell(row=row, column=column_index_from_string('L')).value,
                'specialty': ws_f1.cell(row=row, column=column_index_from_string('M')).value,
                'data_references': ws_f1.cell(row=row, column=column_index_from_string('J')).value,
                'qualification': ws_f1.cell(row=row, column=column_index_from_string('N')).value,
                'gender': list(ws_f1.cell(row=row, column=column_index_from_string('W')).value)[0],
                'SNILS': ws_f1.cell(row=row, column=column_index_from_string('X')).value,
                'on_EPGU': "Да",
                'citizenship': "RU",
                'on_blank': "Да",
                'base_references': "",
                'basis_acceptance': "",
                'email': "",
                'chair_gec': "",
                'previous_document_education': f"Аттестат о среднем образовании, {ws_f1.cell(row=row, column=column_index_from_string('P')).value}",
                'document_view': "",
                'solution_gec': "",
                'term_accumulation': f"{ws_f1.cell(row=row, column=column_index_from_string('R')).value} года",
                'mark': FIO[fio],
                'rektor': "Данилова Оксана Вячеславовна",
            }
            data_student.append(student)

    # Высота строк
    ws.row_dimensions[1].height = 50
    ws.row_dimensions[2].height = 50
    ws.row_dimensions[3].height = 50
    ws.row_dimensions[4].height = 50

    # ЗАголовки таблицы
    headers = [["ФАМИЛИЯ", "ИМЯ",  "ОТЧЕСТВО",
               "ДАТА РОЖДЕНИЯ",	"ГОД ОКОНЧАНИЯ",  "КОД СПЕЦИАЛЬНОСТИ",
               "СПЕЦИАЛЬНОСТЬ",	"ДАТА ВЫДАЧИ",	"Квалификация",	"ПОЛ",	"СНИЛС",
               "НА ЕПГУ",	"Гражданство",	"На бланке",	"Основание выдачи документа",
               "Основание приема на обучение",	"Адрес электронной почты (email)",	"Председатель Государственной экзаменационной комиссии",
               "Предыдущий документ об образовании или об образовании и о квалификации",
               "Вид документа",	"Решение Государственной экзаменационной комиссии",
               "Срок освоения образовательной программы по очной форме обучения"
                ]]

    # Добавление заголовков
    for row in headers:
        ws.append(row)

    # Объединение ячеек
    ws.merge_cells('W1:DM1')
    ws.merge_cells('DO1:DP1')

    # Добавление значений после объеденной фячейки и в
    ws['W1'].value = "Сведения о содержании и результатах освоения образовательной программы среднего профессионального образования [ Наименование учебных предметов, курсов, дисциплин (модулей), практик | Общее количество часов | Оценка ]"
    ws['DN1'].value = "Курсовые работы (проекты) [ Курсовые работы (проекты) | Оценка ]"
    ws['DO1'].value = "Дополнительные сведения [ Содержание дополнительных сведений ]"
    ws['DQ1'].value = "ПОДПИСИ"
    ws['CY3'].value = "Практики"
    ws['CY4'].value = "+в том числе:"
    ws['DL3'].value = "Государственная итоговая аттестация"
    ws['DL4'].value = "+в том числе:"
    ws['DQ3'].value = "Ректор"

    # Назначение ширины столбцов
    for i in range(1, ws.max_column + 1):
        letter = get_column_letter(i)
        ws.column_dimensions[letter].width = 35

    # Выбор заголовков
    INFO_1 = ws['A1:G2']
    INFO_2 = ws['H1:V2']
    INFO_3 = ws['W1']
    INFO_4 = ws['DO1:DP2']
    INFO_5 = ws['DQ1:DQ2']
    INFO_6 = ws['CY3:DK4']
    INFO_7 = ws['DL3:DN4']
    INFO_8 = ws['A3:DQ4']

    thins = Side(border_style="medium", color="211c16")
    double = Side(border_style="medium", color="211c16")

    # Изменение цвета зоголовков
    for row in INFO_1:
        for cell in row:
            cell.fill = PatternFill('solid', fgColor='2e75b6')
            cell.border = Border(top=double, bottom=double, left=thins, right=thins)

    for row in INFO_2:
        for cell in row:
            cell.fill = PatternFill('solid', fgColor='92d050')
            cell.border = Border(top=double, bottom=double, left=thins, right=thins)

    INFO_3.fill = PatternFill('solid', fgColor='ffe699')
    INFO_3.border = Border(top=double, bottom=double, left=thins, right=thins)

    for row in ws['W2:DN2']:
        for cell in row:
            cell.fill = PatternFill('solid', fgColor='ffe699')
            cell.border = Border(top=double, bottom=double, left=thins, right=thins)

    ws['DN1'].fill = PatternFill('solid', fgColor='ffe699')
    ws['DN1'].border = Border(top=double, bottom=double, left=thins, right=thins)

    for row in INFO_4:
        for cell in row:
            cell.fill = PatternFill('solid', fgColor='b4c7e7')
            cell.border = Border(top=double, bottom=double, left=thins, right=thins)

    for row in INFO_5:
        for cell in row:
            cell.fill = PatternFill('solid', fgColor='a7074b')
            cell.border = Border(top=double, bottom=double, left=thins, right=thins)

    for row in INFO_6:
        for cell in row:
            cell.fill = PatternFill('solid', fgColor='ffcccc')
            cell.border = Border(top=double, bottom=double, left=thins, right=thins)

    for row in INFO_7:
        for cell in row:
            cell.fill = PatternFill('solid', fgColor='ffffcc')
            cell.border = Border(top=double, bottom=double, left=thins, right=thins)

    for row in INFO_8:
        for cell in row:
            cell.border = Border(top=double, bottom=double, left=thins, right=thins)

    # Фильтрация данных из массива Subject по предметам (sub), практикам (theme_praktik), курсовым работам (name_kurs_job)
    sub = [j for j in Subject if (j != 'в том числе:' and 'аттестация' not in j
                and 'экзамен' not in j and 'рактика' not in j and
                'курсовая' not in j)]
    theme_praktik = [j for j in Subject if ('рактика' in j and j != "Практика")]
    theme_praktik.append(theme_praktik.pop(0))
    name_kurs_job = [j for j in Subject if ('курсовая'in j)]

    # Извлечение индексов ищ массива Subject по предметам (sub), практикам (theme_praktik), курсовым работам (name_kurs_job)
    sub_index = [j for j in range(0, len(Subject)) if (Subject[j] != 'в том числе:' and 'аттестация' not in Subject[j]
                    and 'экзамен' not in Subject[j] and 'рактика' not in Subject[j] and
                    'курсовая' not in Subject[j])]
    sub_hour = [Subjetc_hour[j] for j in range(0, len(Subject)) if (Subject[j] != 'в том числе:' and 'аттестация' not in Subject[j]
                    and 'экзамен' not in Subject[j] and 'рактика' not in Subject[j] and
                    'курсовая' not in Subject[j])]

    index_praktiks = [j for j in range(0, len(Subject)) if ('рактика' in Subject[j] and Subject[j] != "Практика")]
    index_praktiks.append(index_praktiks.pop(0))

    index_kurs_job = [j for j in range(0, len(Subject)) if ('курсовая'in Subject[j])]
    all_praktik_time = [j for j in range(0, len(Subject)) if (Subject[j] == "Практика")]
    gia = [j for j in range(0, len(Subject)) if ('аттестация' in Subject[j])]
    dem_dip = [j for j in range(0, len(Subject)) if ('экзамен' in Subject[j])]

    # Занесение предметов, практик и курсовых работ в столбцы заголовков
    for i in range(0, len(sub)):
        ws.cell(row=3, column=column_index_from_string('W') + i).value = sub[i]

    for i in range(0, len(theme_praktik)):
        ws.cell(row=4, column=column_index_from_string('CZ') + i).value = theme_praktik[i]

    ws['DN3'].value = name_kurs_job[0]

    # Занесение данных о студентах в таблицу
    mark = {
        '+': "зачтено",
        3: 'удолетворительно',
        4: 'хорошо',
        5: 'отлично',
        'х': "x"
    }

    for i in range(0, len(data_student)):
        ws.row_dimensions[i + 5].height = 50

        ws.cell(row=i + 5, column=column_index_from_string('A')).value = data_student[i]['surname']
        ws.cell(row=i + 5, column=column_index_from_string('B')).value = data_student[i]['name']
        ws.cell(row=i + 5, column=column_index_from_string('C')).value = data_student[i]['patronymic']
        ws.cell(row=i + 5, column=column_index_from_string('D')).value = data_student[i]['birthday']
        ws.cell(row=i + 5, column=column_index_from_string('E')).value = data_student[i]['year_references']
        ws.cell(row=i + 5, column=column_index_from_string('F')).value = data_student[i]['speciality_code']
        ws.cell(row=i + 5, column=column_index_from_string('G')).value = data_student[i]['specialty']
        ws.cell(row=i + 5, column=column_index_from_string('H')).value = data_student[i]['data_references']
        ws.cell(row=i + 5, column=column_index_from_string('I')).value = data_student[i]['qualification']
        ws.cell(row=i + 5, column=column_index_from_string('J')).value = data_student[i]['gender']
        ws.cell(row=i + 5, column=column_index_from_string('K')).value = data_student[i]['SNILS']
        ws.cell(row=i + 5, column=column_index_from_string('L')).value = data_student[i]['on_EPGU']
        ws.cell(row=i + 5, column=column_index_from_string('M')).value = data_student[i]['citizenship']
        ws.cell(row=i + 5, column=column_index_from_string('N')).value = data_student[i]['on_blank']
        ws.cell(row=i + 5, column=column_index_from_string('O')).value = data_student[i]['base_references']
        ws.cell(row=i + 5, column=column_index_from_string('P')).value = data_student[i]['basis_acceptance']
        ws.cell(row=i + 5, column=column_index_from_string('Q')).value = data_student[i]['email']
        ws.cell(row=i + 5, column=column_index_from_string('R')).value = data_student[i]['chair_gec']
        ws.cell(row=i + 5, column=column_index_from_string('S')).value = data_student[i]['previous_document_education']
        ws.cell(row=i + 5, column=column_index_from_string('T')).value = data_student[i]['document_view']
        ws.cell(row=i + 5, column=column_index_from_string('U')).value = data_student[i]['solution_gec']
        ws.cell(row=i + 5, column=column_index_from_string('V')).value = data_student[i]['term_accumulation']

        ws.cell(row=i + 5, column=column_index_from_string('A')).border = Border(top=double, bottom=double, left=thins, right=thins)
        ws.cell(row=i + 5, column=column_index_from_string('B')).border = Border(top=double, bottom=double, left=thins, right=thins)
        ws.cell(row=i + 5, column=column_index_from_string('C')).border = Border(top=double, bottom=double, left=thins, right=thins)
        ws.cell(row=i + 5, column=column_index_from_string('D')).border = Border(top=double, bottom=double, left=thins, right=thins)
        ws.cell(row=i + 5, column=column_index_from_string('E')).border = Border(top=double, bottom=double, left=thins, right=thins)
        ws.cell(row=i + 5, column=column_index_from_string('F')).border = Border(top=double, bottom=double, left=thins, right=thins)
        ws.cell(row=i + 5, column=column_index_from_string('G')).border = Border(top=double, bottom=double, left=thins, right=thins)
        ws.cell(row=i + 5, column=column_index_from_string('H')).border = Border(top=double, bottom=double, left=thins, right=thins)
        ws.cell(row=i + 5, column=column_index_from_string('I')).border = Border(top=double, bottom=double, left=thins, right=thins)
        ws.cell(row=i + 5, column=column_index_from_string('J')).border = Border(top=double, bottom=double, left=thins, right=thins)
        ws.cell(row=i + 5, column=column_index_from_string('K')).border = Border(top=double, bottom=double, left=thins, right=thins)
        ws.cell(row=i + 5, column=column_index_from_string('L')).border = Border(top=double, bottom=double, left=thins, right=thins)
        ws.cell(row=i + 5, column=column_index_from_string('M')).border = Border(top=double, bottom=double, left=thins, right=thins)
        ws.cell(row=i + 5, column=column_index_from_string('N')).border = Border(top=double, bottom=double, left=thins, right=thins)
        ws.cell(row=i + 5, column=column_index_from_string('O')).border = Border(top=double, bottom=double, left=thins, right=thins)
        ws.cell(row=i + 5, column=column_index_from_string('P')).border = Border(top=double, bottom=double, left=thins, right=thins)
        ws.cell(row=i + 5, column=column_index_from_string('Q')).border = Border(top=double, bottom=double, left=thins, right=thins)
        ws.cell(row=i + 5, column=column_index_from_string('R')).border = Border(top=double, bottom=double, left=thins, right=thins)
        ws.cell(row=i + 5, column=column_index_from_string('S')).border = Border(top=double, bottom=double, left=thins, right=thins)
        ws.cell(row=i + 5, column=column_index_from_string('T')).border = Border(top=double, bottom=double, left=thins, right=thins)
        ws.cell(row=i + 5, column=column_index_from_string('U')).border = Border(top=double, bottom=double, left=thins, right=thins)
        ws.cell(row=i + 5, column=column_index_from_string('V')).border = Border(top=double, bottom=double, left=thins, right=thins)

        for j in range(0, len(sub_index)):
            ws.cell(row=i + 5, column=column_index_from_string('W') + j).value = f"{sub_hour[j]} | {mark[data_student[i]['mark'][sub_index[j]]]}"
            ws.cell(row=i + 5, column=column_index_from_string('W') + j).border = Border(top=double, bottom=double, left=thins, right=thins)

        ws.cell(row=i + 5, column=column_index_from_string('CY')).value = f"{Subjetc_hour[all_praktik_time[0]]} | {mark[data_student[i]['mark'][all_praktik_time[0]]]}"
        ws.cell(row=i + 5, column=column_index_from_string('CY')).border = Border(top=double, bottom=double, left=thins, right=thins)

        for j in range(0, len(index_praktiks)):
            ws.cell(row=i + 5, column=column_index_from_string('CZ') + j).value = f"{Subjetc_hour[index_praktiks[j]]} | {mark[data_student[i]['mark'][index_praktiks[j]]]}"
            ws.cell(row=i + 5, column=column_index_from_string('CZ') + j).border = Border(top=double, bottom=double, left=thins, right=thins)

        ws.cell(row=i + 5, column=column_index_from_string('DL')).value = f"{Subjetc_hour[gia[0]]} | {mark[data_student[i]['mark'][gia[0]]]}"
        ws.cell(row=i + 5, column=column_index_from_string('DM')).value = f"{Subject[dem_dip[0]]} | {Subjetc_hour[dem_dip[0]]} | {mark[data_student[i]['mark'][dem_dip[0]]]}"
        ws.cell(row=i + 5, column=column_index_from_string('DN')).value = f"{mark[data_student[i]['mark'][index_kurs_job[0]]]}"

        ws.cell(row=i + 5, column=column_index_from_string('DL')).border = Border(top=double, bottom=double, left=thins, right=thins)
        ws.cell(row=i + 5, column=column_index_from_string('DM')).border = Border(top=double, bottom=double, left=thins, right=thins)
        ws.cell(row=i + 5, column=column_index_from_string('DN')).border = Border(top=double, bottom=double, left=thins, right=thins)

        ws.cell(row=i + 5, column=column_index_from_string('DP')).value = "Форма обучения: очная"
        ws.cell(row=i + 5, column=column_index_from_string('DP')).fill = PatternFill('solid', fgColor='b4c7e7')
        ws.cell(row=i + 5, column=column_index_from_string('DO')).fill = PatternFill('solid', fgColor='b4c7e7')

        ws.cell(row=i + 5, column=column_index_from_string('DP')).border = Border(top=double, bottom=double, left=thins, right=thins)
        ws.cell(row=i + 5, column=column_index_from_string('DO')).border = Border(top=double, bottom=double, left=thins, right=thins)

        ws.cell(row=i + 5, column=column_index_from_string('DQ')).value = data_student[i]['rektor']
        ws.cell(row=i + 5, column=column_index_from_string('DQ')).border = Border(top=double, bottom=double, left=thins, right=thins)

    path = os.path.join(f'C:\\Users\\{os.getlogin()}\\Documents\\Сведения студентов из ФРДО {spec}.xlsx')
    new_file.save(path)
    return path


if __name__ == "__main__":
    ParcerXlsxData()
